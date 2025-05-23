import pandas as pd
from math import sqrt, log as ln, pi
import os
from openpyxl.styles import Alignment
import metric_to_imperical as mti

# Read in the engine data to a dictionary of dataframes
current_dir = os.path.dirname(__file__)
file_path = os.path.join(current_dir,'ICE Measured Data.xlsx')
excel_file = pd.ExcelFile(file_path)
tests = {'One-Quarter Throttle': None, 'One-Half Throttle': None, 
         'Three-Quarter Throttle': None, 'Full Throttle': None}
for sheet_name in excel_file.sheet_names:
    tests[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
imperical_values = {sheet_name: df.copy() for sheet_name, df in tests.items()}

# Displays the first column and last 10 columns of each dataframe
def display_df(tests):
	for sheet_name, df in tests.items():
		print(sheet_name)
		print(df.iloc[:, [0]].join(df.iloc[:, -5:]))

# Engine constants
connecting_rod_length = 0.105  # m
engine_displacement = 163e-6  # m^3
bore_diameter = 0.068  # m
stroke_length = 0.045  # m
n_p = 2 # Revolutions per power stroke

# Fuel constants
hhv = 47.3e6  # J/kg (higher heating value of gasoline)
density = 750  # kg/m^3

# Measured Brake power (MBP) calculation in Watts
eta = 0.8  # Efficiency of the engine
for test in tests.values():
	test['MBP (W)'] = test['Oil Pressure (Pa)'] * test['Oil Flow Rate (m^3/s)']* eta
# Imperical
for sheet,data in tests.items():
	imperical_values[sheet]['MBP (hp)'] = mti.mbp_conversion(data[['MBP (W)']])

# Fuel flow rate (FFR) calculation in m^3/s
for test in tests.values():
	test['FFR (m^3/s)'] = test['Difference (kg)'] / test['Run Time (s)']
# Imperical
for sheet,data in tests.items():
	imperical_values[sheet]['FFR (GPM)'] = mti.ffr_conversion(data[['FFR (m^3/s)']])

# Brake specific fuel consumption (BSFC) calculation in kg/Nm
for test in tests.values():
	test['BSFC (kg/Nm)'] = test['FFR (m^3/s)'] / test['MBP (W)']
# Imperical
for sheet,data in tests.items():
	imperical_values[sheet]['BSFC (lb/hp-hr)'] = mti.bsfc_conversion(data[['BSFC (kg/Nm)']])

# Rev time calculation in seconds
for test in tests.values():
	test['Rev Time (s)'] = 1 / test['RPM']

# Brake mean effective pressure (BMEP) calculation in Pa
for test in tests.values():
	test['BMEP (Pa)'] = test['MBP (W)'] * n_p / (engine_displacement * test['Rev Time (s)'])
# Imperical
for sheet,data in tests.items():
	imperical_values[sheet]['BMEP (psi)'] = mti.bmep_conversion(data[['BMEP (Pa)']])

# Brake torque calculation in Nm
for test in tests.values():
	test['Brake Torque (Nm)'] = test['MBP (W)'] / (2 * pi * test['RPM'])
# Imperical
for sheet,data in tests.items():	
	imperical_values[sheet]['Brake Torque (lb-ft)'] = mti.torque_conversion(data[['Brake Torque (Nm)']])

# Conversion efficiency (Conv Eff) calculation in %
for test in tests.values():
	test['Conv Eff (%)'] = (test['MBP (W)'] * 100) / (test['FFR (m^3/s)'] * hhv)
# Imperical
for sheet,data in tests.items():
	imperical_values[sheet]['Conv Eff (%)'] = data['Conv Eff (%)']

# Write the data to a results Excel file
def write_results_to_excel(dataframes, output_file_path):
	with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
		for sheet_name, df in dataframes.items():
			df.to_excel(writer, sheet_name=sheet_name, index=False)
			worksheet = writer.sheets[sheet_name]
			for col_idx, col in enumerate(df.columns, start=1):
				column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
				worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = column_width
			# Center align all cells
			for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
				for cell in row:
					cell.alignment = Alignment(horizontal='center', vertical='center')


# Only runs if this file is executed directly
if __name__ == "__main__":
	output_file_path = os.path.join(current_dir, 'Engine Calculations.xlsx')
	write_results_to_excel(tests, output_file_path)
	imp_output_file_path = os.path.join(current_dir, 'Imperical Engine Calculations.xlsx')
	write_results_to_excel(imperical_values, imp_output_file_path)

	# Open the results Excel file
	os.startfile(output_file_path)
	os.startfile(imp_output_file_path)

# print('Tests')
# display_df(tests)
# print('\n')
# print('Imperical Values')
# display_df(imperical_values)
